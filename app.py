import streamlit as st
import pandas as pd
import io

st.set_page_config(page_title="💎 Repricing Tool", layout="wide")

st.title("💎 Repricing Automation (Updated)")

uploaded_file = st.file_uploader("Upload Final File from Project 1", type=["xlsx"])


# ---------------- STEP 1: SIZE GROUP ----------------
def get_size_group(cts):
    try:
        cts = float(cts)
    except:
        return ""

    if cts < 0.30:
        return "<0.30"
    elif 0.30 <= cts <= 0.39:
        return "0.30 - 0.39"
    elif 0.40 <= cts <= 0.49:
        return "0.40 - 0.49"
    elif 0.50 <= cts <= 0.59:
        return "0.50 - 0.59"
    elif 0.60 <= cts <= 0.69:
        return "0.60 - 0.69"
    elif 0.70 <= cts <= 0.79:
        return "0.70 - 0.79"
    elif 0.80 <= cts <= 0.89:
        return "0.80 - 0.89"
    elif 0.90 <= cts <= 0.99:
        return "0.90 - 0.99"

    elif 1.00 <= cts <= 1.10:
        return "1.00 - 1.10"
    elif 1.11 <= cts <= 1.49:
        return "1.11 - 1.49"
    elif 1.50 <= cts <= 1.59:
        return "1.50 - 1.59"
    elif 1.60 <= cts <= 1.99:
        return "1.60 - 1.99"

    elif 2.00 <= cts <= 2.10:
        return "2.00 - 2.10"
    elif 2.11 <= cts <= 2.49:
        return "2.11 - 2.49"
    elif 2.50 <= cts <= 2.59:
        return "2.50 - 2.59"
    elif 2.60 <= cts <= 2.99:
        return "2.60 - 2.99"

    elif 3.00 <= cts <= 3.10:
        return "3.00 - 3.10"
    elif 3.11 <= cts <= 3.49:
        return "3.11 - 3.49"
    elif 3.50 <= cts <= 3.59:
        return "3.50 - 3.59"
    elif 3.60 <= cts <= 3.99:
        return "3.60 - 3.99"

    elif 4.00 <= cts <= 4.10:
        return "4.00 - 4.10"
    elif 4.11 <= cts <= 4.49:
        return "4.11 - 4.49"
    elif 4.50 <= cts <= 4.59:
        return "4.50 - 4.59"
    elif 4.60 <= cts <= 4.99:
        return "4.60 - 4.99"

    elif 5.00 <= cts <= 5.49:
        return "5.00 - 5.49"
    elif 5.50 <= cts <= 5.99:
        return "5.50 - 5.99"

    elif 6 <= cts < 25:
        lower = int(cts)
        upper = lower + 0.99
        return f"{lower:.2f} - {upper:.2f}"

    else:
        return "25+"


# ---------------- NEW STEP: COLOR FILTER ----------------
def filter_color(df):
    df["Color"] = df["Color"].astype(str).str.strip().str.upper()

    allowed_colors = ["D", "E", "F", "G", "H", "I", "J", "K"]

    return df[df["Color"].isin(allowed_colors)]


# ---------------- MAIN ----------------
if uploaded_file:

    df = pd.read_excel(uploaded_file)
    df.columns = df.columns.str.strip()

    st.subheader("📊 Original Data")
    st.dataframe(df.head())

    if st.button("🚀 Generate Repricing File"):

        # STEP 1 → Size Group
        if "Cts." not in df.columns:
            st.error("❌ 'Cts.' column missing")
            st.stop()

        df["Size Group"] = df["Cts."].apply(get_size_group)

        cols = list(df.columns)
        size_col = cols.pop(cols.index("Size Group"))
        cts_index = cols.index("Cts.")
        cols.insert(cts_index + 1, size_col)
        df = df[cols]

        # 🔥 STEP 2 → COLOR FILTER
        if "Color" not in df.columns:
            st.error("❌ 'Color' column missing")
            st.stop()

        df = filter_color(df)

        # STEP 3 → Add Pricing Columns
        if "Cost / Cts." not in df.columns:
            st.error("❌ 'Cost / Cts.' column missing")
            st.stop()

        df["Updated Price"] = ""
        df["Difference"] = ""

        # Excel formula export
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")

            sheet = writer.sheets["Sheet1"]

            headers = [cell.value for cell in sheet[1]]

            from openpyxl.utils import get_column_letter

            cost_col = get_column_letter(headers.index("Cost / Cts.") + 1)
            updated_col = get_column_letter(headers.index("Updated Price") + 1)
            diff_col = get_column_letter(headers.index("Difference") + 1)

            for row in range(2, len(df) + 2):
                formula = f'=IF({updated_col}{row}="","",-ROUND(({cost_col}{row}-{updated_col}{row})/{cost_col}{row},2))'
                sheet[f"{diff_col}{row}"] = formula

        st.success("✅ File Ready (Color Filter + Formula Applied)")

        st.download_button(
            "📥 Download Repricing File",
            data=output.getvalue(),
            file_name="repricing_ready.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )